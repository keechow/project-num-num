"""
Name: 02_PG_chk_nxt_draw.py
Objectives: [PLAYGROUND][4D]Check the relationship btwn current draw and next draw
			criteria: n_range, n_cat, n_csq, n_sum, odd or even
Author: The Echo Telion Project
		echo.telion@gmail.com
"""
import num_analysis as na
from num_analysis import check_num_cat
from num_analysis import identify_n_range
from num_analysis import identify_n_sum
from num_analysis import identify_odd_even
from num_analysis import check_num_csq

from file_handler import read
from openpyxl import load_workbook

def determine_hi_lo(p_dict):
# example of input: {'24': 38, '25': 41, '26': 48, '27': 38, '...}
	list_10x10 = []# make a 10 X 10 list
	for i in range(10):
		list_i = []
		for j in range(10):
			list_i.append(0)
		list_10x10.append(list_i)
	dict_keys = p_dict.keys()
	for each_key in dict_keys:
		i = int(each_key[0])
		j = int(each_key[1])
		value = p_dict[each_key]
		list_10x10[i][j] = value
	return list_10x10

def extract_p_num(selection):
	# return a list containing P123 num, P1 num, P2 num, P3 num
	# selection = 0 : read historical 4D.txt
	# selection = 1 : read new 4D.txt

	if selection == 0:
		result = read("4D.txt")
	else:
		result = read("4D-latest.txt")
	p123 = []    #latest prize 1,2,3
	p1_only = []
	p2_only = []
	p3_only = []

	for each_draw in result:
		# filter data stream from result file into different p_cat
		p123.append([each_draw[0],each_draw[2],each_draw[3],each_draw[4]])
		p1_only.append(each_draw[2])
		p2_only.append(each_draw[3])
		p3_only.append(each_draw[4])
	return [p123,p1_only,p2_only,p3_only]

def p1_num(selection):
	# selection = 0 : read historical 4D.txt
	# selection = 1 : read new 4D.txt
	return extract_p_num(selection)[1]

def p2_num(selection):
	# selection = 0 : read historical 4D.txt
	# selection = 1 : read new 4D.txt
	return extract_p_num(selection)[2]

def p3_num(selection):
	# selection = 0 : read historical 4D.txt
	# selection = 1 : read new 4D.txt
	return extract_p_num(selection)[3]

def p123_num(selection):
	# selection = 0 : read historical 4D.txt
	# selection = 1 : read new 4D.txt
	return extract_p_num(selection)[0]

def p_dict_print(p_dict):
	for outer in range(10):
		counter = 0
		for inner in range(10):
			keys = str(outer)+str(inner)
			print keys,": ", p_dict[keys]
			counter += p_dict[keys]
		print "Total = ", counter
		print "Average = ", (counter / 10.0)
		print
		print

def cnr_nncat_printer(result):
	result_key_list = result.keys()
		# result_key_list = [1i24, 9i12, 8i1, 5i6,....]
	cnr_list = []
	for each in range(10):
		cnr_list.append([])
	for each in cnr_list:
		each.append(0)
	for each in result_key_list:
		cnr = int(each[0])
		cnr_list[cnr][0] += result[each]
	for each in result_key_list:
		cnr = int(each[0])
		total = cnr_list[cnr][0]
		average = str((result[each] / float(total))*100)
		cnr_list[cnr].append(str(each) + " : " + str(result[each]) + " (Pct: " + average + ")")




	for each in cnr_list:
		print each
		print
		print

def check_cnr_nnr(p_result_list):
	return_p_dict = {}
    # create a dict to count cnr vs nnr
	for i in range(10):
		for j in range(10):
			key = str(i)+str(j)
			return_p_dict[key] = 0
	p_counter = 0    # use for looping the p_result_list
	while p_counter < (len(p_result_list) - 1):
		cnr = str(p_result_list[p_counter][0])
		nnr = str(p_result_list[p_counter + 1][0])
		key = cnr + nnr
		if key in return_p_dict:
			return_p_dict[key] += 1
		p_counter += 1
	return return_p_dict

def check_cnr_nnr_allprize(p_result_list):
	return_p1_dict = {}
	return_p2_dict = {}
	return_p3_dict = {}
    # create a dict to count cnr vs nnr
	for i in range(10):
		for j in range(10):
			key = str(i)+str(j)
			return_p1_dict[key] = 0
			return_p2_dict[key] = 0
			return_p3_dict[key] = 0
	p_counter = 0    # use for looping the p_result_list

	while p_counter < (len(p_result_list) - 1):
		cnr_p1 = str(p_result_list[p_counter][1][0])
		cnr_p2 = str(p_result_list[p_counter][2][0])
		cnr_p3 = str(p_result_list[p_counter][3][0])

		nnr_p1 = str(p_result_list[p_counter + 1][1][0])
		nnr_p2 = str(p_result_list[p_counter + 1][2][0])
		nnr_p3 = str(p_result_list[p_counter + 1][3][0])

		key_cp1_np1 = cnr_p1 + nnr_p1    #current p1 num_range vs next p1 num range
		key_cp1_np2 = cnr_p1 + nnr_p2    #current p1 num_range vs next p2 num range
		key_cp1_np3 = cnr_p1 + nnr_p3    #current p1 num_range vs next p3 num range
		key_cp1_list = [key_cp1_np1, key_cp1_np2, key_cp1_np3]

		key_cp2_np1 = cnr_p2 + nnr_p1    #current p2 num_range vs next p1 num range
		key_cp2_np2 = cnr_p2 + nnr_p2    #current p2 num_range vs next p2 num range
		key_cp2_np3 = cnr_p2 + nnr_p3    #current p2 num_range vs next p3 num range
		key_cp2_list = [key_cp2_np1, key_cp2_np2, key_cp2_np3]

		key_cp3_np1 = cnr_p3 + nnr_p1    #current p3 num_range vs next p1 num range
		key_cp3_np2 = cnr_p3 + nnr_p2    #current p3 num_range vs next p2 num range
		key_cp3_np3 = cnr_p3 + nnr_p3    #current p3 num_range vs next p3 num range
		key_cp3_list = [key_cp3_np1, key_cp3_np2, key_cp3_np3]

		for each in key_cp1_list:
			if each in return_p1_dict:
				return_p1_dict[each] += 1
		for each in key_cp2_list:
			if each in return_p2_dict:
				return_p2_dict[each] += 1
		for each in key_cp3_list:
			if each in return_p3_dict:
				return_p3_dict[each] +=1

		p_counter += 1
	return [return_p1_dict, return_p2_dict, return_p3_dict]

def check_cnr_nnr_otherprize(p_result_list1, p_result_list2):
	#objective: comparing cnr with nnr, return count (dict)
	return_p_dict = {}
    # create a dict to count cnr vs nnr
	for i in range(10):
		for j in range(10):
			key = str(i)+str(j)
			return_p_dict[key] = 0
	p_counter = 0    # use for looping the p_result_list
	while p_counter < (len(p_result_list1) - 1):
		cnr = str(p_result_list1[p_counter][0])
		nnr = str(p_result_list2[p_counter + 1][0])
		key = cnr + nnr
		if key in return_p_dict:
			return_p_dict[key] += 1
		p_counter += 1
	return return_p_dict

def check_cnr_nncat_single_p_cat(p_result_list):

	#1) create 1 dictionary to count cnr vs nncat
	#2) determine cnr from result list
	#3) determine nncat from result list
	#4) populate
	#5) i24 = 0.1984, i12 = 0.2314, i6 = 3.7037, i4 = 2.7777, i1 = 100

	return_p_dict = {}
		# create a dict to count cnr vs nncat
	n_cat_value = {"i24":0.1984, "i12":0.2314, "i6":3.7037, "i4":2.7777, "i1":100}

	for i in range(10):
		for each in ["i24", "i12", "i6", "i4", "i1"]:
			key = str(i) + each
			return_p_dict[key] = 0

	p_counter = 0    # use for looping the p_result_list
	while p_counter < (len(p_result_list) - 1):
		cnr = str(p_result_list[p_counter][0])
		next_num = (p_result_list[p_counter + 1])
		next_num_cat = check_num_cat(next_num)
		nn_cat_key = "i" + str(next_num_cat)
		nn_cat_key_value = n_cat_value[nn_cat_key]
		key = cnr + nn_cat_key
		if key in return_p_dict:
			return_p_dict[key] += 1
		p_counter += 1
	return return_p_dict

def check_cnr_nncat_diff_p_cat(p_result_list, c_p_cat, nxt_p_cat):

	#1) create 1 dictionary to count cnr vs nncat
	#2) determine cnr from result list
	#3) determine nncat from result list
	#4) populate
	#5) i24 = 0.1984, i12 = 0.2314, i6 = 3.7037, i4 = 2.7777, i1 = 100

	return_p_dict = {}
		# create a dict to count cnr vs nncat
	n_cat_value = {"i24":0.1984, "i12":0.2314, "i6":3.7037, "i4":2.7777, "i1":100}

	for i in range(10):
		for each in ["i24", "i12", "i6", "i4", "i1"]:
			key = str(i) + each
			return_p_dict[key] = 0

	# ['439516', '4030', '4304', '7899'], ['439616', '8414', '8575', '8339'],
	p_counter = 0    # use for looping the p_result_list
	while p_counter < (len(p_result_list) - 1):
		cnr = str(p_result_list[p_counter][c_p_cat][0])
		next_num = (p_result_list[p_counter + 1][nxt_p_cat])
		next_num_cat = check_num_cat(next_num)
		nn_cat_key = "i" + str(next_num_cat)
		nn_cat_key_value = n_cat_value[nn_cat_key]
		key = cnr + nn_cat_key
		if key in return_p_dict:
			return_p_dict[key] += 1
		p_counter += 1
	return return_p_dict

def check_cnr_nncat_p123(p_result_list, c_p_cat):

	#1) create 1 dictionary to count cnr vs nncat
	#2) determine cnr from result list
	#3) determine nncat from result list
	#4) populate
	#5) i24 = 0.1984, i12 = 0.2314, i6 = 3.7037, i4 = 2.7777, i1 = 100

	return_p_dict = {}
		# create a dict to count cnr vs nncat
	n_cat_value = {"i24":0.1984, "i12":0.2314, "i6":3.7037, "i4":2.7777, "i1":100}

	for i in range(10):
		for each in ["i24", "i12", "i6", "i4", "i1"]:
			key = str(i) + each
			return_p_dict[key] = 0

	# ['439516', '4030', '4304', '7899'], ['439616', '8414', '8575', '8339'],
	p_counter = 0    # use for looping the p_result_list
	while p_counter < (len(p_result_list) - 1):
		cnr = str(p_result_list[p_counter][c_p_cat][0])
		next_num = (p_result_list[p_counter + 1][1:])
		for each in next_num:
			next_num_cat = check_num_cat(each)
			nn_cat_key = "i" + str(next_num_cat)
			#nn_cat_key_value = n_cat_value[nn_cat_key]
			key = cnr + nn_cat_key
			if key in return_p_dict:
				return_p_dict[key] += 1
		p_counter += 1
	return return_p_dict

def check_cnr_nnsum_multiple_p_cat(p_result_list, c_p_cat, nxt_p_cat):
	# 1. For 0000 ~ 9999, digit sum is from 0 to 36
	# 2. Create a list to store cnr 0-9
	# 3. For each element in cnr_list, attach a list of 36 int 0
	# 4. This is used to count nn_sum
	# 3. Read cnr and get nn_sum using na.calc_digit_sum()
	# 4. Populate return_dict with nn_sum

	return_list = []
	for i in range(10):
		nnr = []
		for j in range(37):
			nnr.append([0])
		return_list.append(nnr)

	# ['439516', '4030', '4304', '7899'], ['439616', '8414', '8575', '8339'],
	p_counter = 0    # use for looping the p_result_list
	while p_counter < (len(p_result_list) - 1):
		cnr_idx = int(p_result_list[p_counter][c_p_cat][0])    # we want to num range as index

		nxt_num = int(p_result_list[p_counter + 1][nxt_p_cat])

		nn_sum_idx = na.calc_digit_sum(int(nxt_num))

		return_list[cnr_idx][nn_sum_idx][0] +=1
		    # cnr_idx and nn_sum_idx is the index for return_list
		    # count the occurrences of nn_sum for a given cnr
		p_counter += 1
	return return_list

def check_cnr_nnsum_p123(p_result_list, c_p_cat):
	# This is to count nnsum for all p123
	# 1. For 0000 ~ 9999, digit sum is from 0 to 36
	# 2. Create a list to store cnr 0-9
	# 3. For each element in cnr_list, attach a list of 36 int 0
	# 4. This is used to count nn_sum
	# 3. Read cnr and get nn_sum using na.calc_digit_sum()
	# 4. Populate return_dict with nn_sum

	return_list = []
	for i in range(10):
		nnr = []
		for j in range(37):
			nnr.append([0])
		return_list.append(nnr)

	# ['439516', '4030', '4304', '7899'], ['439616', '8414', '8575', '8339'],
	p_counter = 0    # use for looping the p_result_list
	while p_counter < (len(p_result_list) - 1):
		cnr_idx = int(p_result_list[p_counter][c_p_cat][0])    # we want to num range as index

		nxt_num = p_result_list[p_counter + 1][1:]

		for each in nxt_num:
			nn_sum_idx = na.calc_digit_sum(int(each))
			return_list[cnr_idx][nn_sum_idx][0] +=1
		    # cnr_idx and nn_sum_idx is the index for return_list
		    # count the occurrences of nn_sum for a given cnr
		p_counter += 1
	return return_list

def cnr_nnsum_port2excel(result_list, excel_file_name, worksheet_name):
	# module to take string data output and port into excel
	# 1. Read from result list
	# 2. Filter our unwanted string
	# 3. Load excel file
	# 4. Open sheet, locate cell for each data, populate

    # result list = 10 elements type list, each result list sub_element consists of 36 lists, containing the count for each occurrence

	workbook = load_workbook(excel_file_name)    # load workbook

	worksheet = workbook.create_sheet(title=worksheet_name)    # load worksheet

	# set the title/desc for row 1
	cell_count = 1
	while cell_count < 38:
		worksheet.cell(row = 1, column = cell_count).value = str(cell_count-1)
		cell_count += 1
    # set title for column 1
	cell_ctr = 1
	while cell_ctr < 12:
		worksheet.cell(row=cell_ctr, column=1).value = str(cell_ctr-1)
		cell_ctr += 1

	#populating worksheet with data
	row_count = 2
	column_count = 2

	while row_count < 12:
		while column_count < 38:
			value = str(result_list[row_count-2][column_count-2])
			value1 = value[1:-1]    #delete bracket [] for list to be entered into xls cell
			worksheet.cell(row=row_count, column=column_count).value = int(value1)
			column_count += 1
		column_count = 2
		row_count +=1

	workbook.save(excel_file_name)

def cnr_nnsum_test_module(result_list):
	# result: ['439516', '4030', '4304', '7899'], ['439616', '8414', '8575', '8339']
	# no_list: ['9321', '9331', '111231', '111232']

	print "Result List length: ", len(result_list)

	cnr_nnsum_list = []
	counter = 0

	while counter < len(result_list)-1:    #generating a list of cnr_nnsum for result list
		cp_cat = 1
		while cp_cat < 4:
			np_cat = 1
			result_cnr = result_list[counter][cp_cat][0]
			while np_cat < 4:
				result_nnsum = str(na.calc_digit_sum(int(result_list[counter + 1][np_cat])))
				cnr_nnsum_element = result_cnr + str(cp_cat) + str(np_cat) + result_nnsum
				cnr_nnsum_list.append(cnr_nnsum_element)
				np_cat += 1
			cp_cat += 1
		counter += 1

	print "cnr_nnsum_list len: ", len(cnr_nnsum_list)

	# creating no_list
	no_list = []
	for each in read("cnr_nnsum_no_list.txt"):
		no_list.append(each[0])

	print "no list len: ", len(no_list)

	# match cnr_nnsum_list with no_list. append matched element into return list
	return_list = []
	for each in cnr_nnsum_list:		#comparing elements in cnr_nnsum_list with no_list
		if each in no_list:
			return_list.append(each)
	return return_list

	print "===================="
	print "=====COMPLETED======"
	print "===================="










# CNR vs NN_SUM


"""
cpcat1_npcat1 = check_cnr_nnsum_multiple_p_cat(result,1,1)
cpcat1_npcat2 = check_cnr_nnsum_multiple_p_cat(result,1,2)
cpcat1_npcat3 = check_cnr_nnsum_multiple_p_cat(result,1,3)
cpcat2_npcat1 = check_cnr_nnsum_multiple_p_cat(result,2,1)
cpcat2_npcat2 = check_cnr_nnsum_multiple_p_cat(result,2,2)
cpcat2_npcat3 = check_cnr_nnsum_multiple_p_cat(result,2,3)
cpcat3_npcat1 = check_cnr_nnsum_multiple_p_cat(result,3,1)
cpcat3_npcat2 = check_cnr_nnsum_multiple_p_cat(result,3,2)
cpcat3_npcat3 = check_cnr_nnsum_multiple_p_cat(result,3,3)

cpcat1_npcat123 =  check_cnr_nnsum_p123(p123_num(),1)
cpcat2_npcat123 =  check_cnr_nnsum_p123(p123_num(),2)
cpcat3_npcat123 =  check_cnr_nnsum_p123(p123_num(),3)
"""

"""
xlsx_file = "result_cnr_nnsum.xlsx"
cnr_nnsum_port2excel(cpcat1_npcat1, xlsx_file, "cpcat1_npcat1")
cnr_nnsum_port2excel(cpcat1_npcat2, xlsx_file, "cpcat1_npcat2")
cnr_nnsum_port2excel(cpcat1_npcat3, xlsx_file, "cpcat1_npcat3")
cnr_nnsum_port2excel(cpcat1_npcat123, xlsx_file, "cpcat1_npcat123")

cnr_nnsum_port2excel(cpcat2_npcat1, xlsx_file, "cpcat2_npcat1")
cnr_nnsum_port2excel(cpcat2_npcat2, xlsx_file, "cpcat2_npcat2")
cnr_nnsum_port2excel(cpcat2_npcat3, xlsx_file, "cpcat2_npcat3")
cnr_nnsum_port2excel(cpcat2_npcat123, xlsx_file, "cpcat2_npcat123")

cnr_nnsum_port2excel(cpcat3_npcat1, xlsx_file, "cpcat3_npcat1")
cnr_nnsum_port2excel(cpcat3_npcat2, xlsx_file, "cpcat3_npcat2")
cnr_nnsum_port2excel(cpcat3_npcat3, xlsx_file, "cpcat3_npcat3")
cnr_nnsum_port2excel(cpcat3_npcat123, xlsx_file, "cpcat3_npcat123")
"""

#result_cnr_nnsum.xlsx

"""

# CNR vs NN_cat for P1 only
p1_only = p1_num(0)

cnr_nncat_p1 = check_cnr_nncat_single_p_cat(p1_only)
cnr_nncat_p1_p1 = check_cnr_nncat_diff_p_cat(p123_num(),3,3)
cnr_nncat_p1_p123 = check_cnr_nncat_multiple_p_cat(p123_num(),3)

#cnr_nncat_printer(cnr_nncat_p1_p123)
cnr_nncat_printer((cnr_nncat_p1_p1))
#cnr_nncat_printer(cnr_nncat_p1_p123)


"""
# this is to check cnr vs nnr: all p_cat

p1_only = p1_num(1) #setting up input for comparison module
p2_only = p2_num(1)
p3_only = p3_num(1)
p123 = p123_num(1)
p1_dict = check_cnr_nnr(p1_only)
p2_dict = check_cnr_nnr(p2_only)
p3_dict = check_cnr_nnr(p3_only)
p123_dict = check_cnr_nnr_allprize(p123)
p1_v_p2_dict = check_cnr_nnr_otherprize(p1_only, p2_only)
p1_v_p3_dict = check_cnr_nnr_otherprize(p1_only, p3_only)
p2_v_p1_dict = check_cnr_nnr_otherprize(p2_only, p1_only)
p2_v_p3_dict = check_cnr_nnr_otherprize(p2_only, p3_only)
p3_v_p1_dict = check_cnr_nnr_otherprize(p3_only, p1_only)
p3_v_p2_dict = check_cnr_nnr_otherprize(p3_only, p2_only)

p_dict_print(p1_dict)



"""
#

prize1 = []
prize2 = []
prize3 = []

for each_draw in result:
	year = each_draw[0][-2:]
	p1 = each_draw[2]
	p2 = each_draw[3]
	p3 = each_draw[4]
	if year == "14" or year == "15" or year == "16":
		prize1.append(p1)
		prize2.append(p2)
		prize3.append(p3)
prize123 = prize1 + prize2 + prize3

# we want to eliminate same element within prize123
prize123_clean = na.clean_duplicate(prize123)
eliminated_num_list = na.clean_duplicate(i1_num + i4_num + prize123_clean) #nums that we don't want

clean_4d_num_list = []
for each in all_4D_num:
	if each not in eliminated_num_list:
		clean_4d_num_list.append(each)

print "prize123_clean len: ", len(prize123_clean)
print "i1num len: ", len(i1_num)
print "i4num len: ", len(i4_num)
print "clean 4d: ", len(clean_4d_num_list)
print "el list: ", len(eliminated_num_list)

"""
print
print "======================================================================"
print "============================== COMPLETED ============================="
print "======================================================================"

