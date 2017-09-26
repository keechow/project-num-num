__author__ = 'keechow'

from  get_result_frm_webpage import build_url_search_term
from  get_result_frm_webpage import parse_data_from_url
from openpyxl import load_workbook
from openpyxl import Workbook

import num_analysis

import json

## save list_data_whole_range into  file for future use
def save_to_file(data, file_name):
	with open(file_name,"w") as outputfile:
		json.dump(data, outputfile)

def load_from_file(file_name):
	return_list_data = []
	with open(file_name) as inputfile:
		json_load_from_file = json.load(inputfile)
		for each_draw in json_load_from_file:
			str_data = [s.encode('utf-8') for s in each_draw]
			return_list_data.append(str_data)
	return return_list_data

def built_url_parse_data(s_month, s_year, num_month, company):
	str_data_collection_start_month = s_month
	str_data_collection_start_year = start_year
	str_data_collection_num_of_month = num_month
	str_data_collection_company = company

	list_search_url = build_url_search_term(str_data_collection_start_month, str_data_collection_start_year, str_data_collection_num_of_month, str_data_collection_company)
		## containing a list of URL to be used to request data from Internet

	list_data_whole_range = []
		## list for containing all parsed data

	for each_month_data_url in list_search_url:
		print each_month_data_url
		monthly_data = parse_data_from_url(each_month_data_url)
		for each_draw_data in monthly_data:
			list_data_whole_range.append(each_draw_data)

	return list_data_whole_range

def nr_3drawSet(list_data, p_cat0, p_cat1, p_cat3):
	## 1. get data from input
	## 2. Extract num range for 3 consecutive draws and save into a list
	## 3. Return the list
	max_loop_count = len(list_data) - 3
	counter = 0

	list_p1_nr_3drawSet = []

	while counter <= max_loop_count:
		num_range1 = list_data[counter][p_cat0][0]
		num_range2 = list_data[counter + 1][p_cat1][0]
		num_range3 = list_data[counter + 2][p_cat3][0]

		list_p1_nr_3drawSet.append(num_range1+num_range2+num_range3)
		counter += 1

	return list_p1_nr_3drawSet

def nr_drawSet(list_data, list_pcat):
	# params: list_data
	# params: list_pcat - type list, containing pcat for num range

	num_of_draw = len(list_pcat)
	max_loop_count = len(list_data) - num_of_draw
	loop_count = 0
	return_list_nr_drawset = []

	while loop_count <= max_loop_count:
		drawset_pattern = ""
		pcat_count = 0
		while pcat_count < num_of_draw:
			pcat = list_pcat[pcat_count]
			pattern = list_data[loop_count + pcat_count][pcat][0]
			drawset_pattern += pattern
			pcat_count += 1
		return_list_nr_drawset.append(drawset_pattern)
		loop_count +=  1

	return return_list_nr_drawset

def nr_4drawSet(list_data, p_cat0, p_cat1, p_cat2, p_cat3):
	## 1. Get data from input
	## 2. Extract num range for 4 consecutive draws and save into a list
	## 3. Return the list

	max_loop_count = len(list_data) - 4
	counter = 0

	list_p1_nr_4drawSet = []

	while counter <= max_loop_count:
		num_range1 = list_data[counter][p_cat0][0]
		num_range2 = list_data[counter + 1][p_cat1][0]
		num_range3 = list_data[counter + 2][p_cat2][0]
		num_range4 = list_data[counter + 3][p_cat3][0]
		list_p1_nr_4drawSet.append(num_range1+num_range2+num_range3+num_range4)
		counter += 1

	return list_p1_nr_4drawSet

def counter_list_3draw(drawSet_data):
	# params: list of data containing draw set pattern
	# count each draw set pattern
	# return a list containing count for each draw set pattern in respective position
	# e.g. element [203] correspond to draw set pattern 203
	return_counter_list = []
	for each in range(1000):
		return_counter_list.append(0)

	for each in drawSet_data:
		int_each = int(each)
		count = return_counter_list[int_each] + 1
		return_counter_list[int_each] = count

	return return_counter_list

def port_counter_list_to_xls(list_counter, s_name):
	# params: counter list
	# objective: port the counter list into xls spreadsheet for easy visualization and further studies

	#1. load xls workbook
	my_wb =  load_workbook("list_3drawset_pattern.xlsx")

	#2. load worksheet in workbook
	sheet_names = my_wb.sheetnames

	if s_name in sheet_names:
		my_ws = my_wb.get_sheet_by_name(s_name)
	else:
		my_ws = my_wb.create_sheet(title=s_name)

	drawset_counter = 0
	while drawset_counter < len(list_counter):
	#3. set initial cell address to port in data
		current_row = drawset_counter + 1
		drawset__column = 1
		count_column = 2
		str_drawset_pattern = str(drawset_counter)
		while (len(str_drawset_pattern) < 3):
			str_drawset_pattern = "0" + str_drawset_pattern
		str_drawset_count = list_counter[drawset_counter]

		my_ws.cell(row = current_row, column = drawset__column).value = str_drawset_pattern
		my_ws.cell(row = current_row, column = count_column).value = str_drawset_count
		drawset_counter += 1
	my_wb.save("list_3drawset_pattern.xlsx")

def remove_duplicate (list_data):
	# removes duplicates in list_data
	clean_list = []
	for each in list_data:
		if not each in clean_list:
			clean_list.append(each)
	return(sorted(clean_list))

def identify_zero_count_pattern(list_data):
	# return 3 draw set pattern that has zero occurence
	list_1K_pattern = []
	list_zero_count_pattern = []
	for each_num in range(1000):
		str_num = str(each_num)
		while (len(str_num) < 3):
			str_num = "0" + str_num
		list_1K_pattern.append(str_num)

	for each in list_1K_pattern:
		if not each in list_data:
			list_zero_count_pattern.append(each)

	return list_zero_count_pattern

def nr_vs_pcat_1_3drawset_count_p2excel(list_data):
	# params: data list from txt file
	# 1) extract 3 draw set pattern (nr vs pcat)
	# 2) count occurrence for each 3draw set pattern
	# 3) port the counter list to excel for future analysis
	# 4) pcat pattern = 111,121,131,  211,221,231, 311,321,331

	for pcat0 in range(1,4):
		for pcat1 in range(1,4):
			nr_3drawset_pattern = nr_3drawSet(list_data, pcat0, pcat1, 1)
			nr_counter_list = counter_list_3draw(nr_3drawset_pattern)
			sheet_name = str(pcat0) + str(pcat1) + "1"
			port_counter_list_to_xls(nr_counter_list, sheet_name)



# 1) get data from the web
# 2) parse raw data
# 3) save data into file
# 4) read data and extract num-range
# 5) extract num-range for 3 consecutive draws
# 6) start counting
#		6a) 3 draws = 1000 variations, 000 ~ 999
#		7b)


##########################################
##                                  RUN  CODE                                 ##
##########################################

"""
### Run once to get data from Internet and save into file ###

start_month = "01"
start_year = "2011"
num_of_month = "81"
company = "magnum"

data_file_name = company + "_" + start_year + "_" + start_month + "_" + num_of_month + "_m.txt"

list_data_whole_range = built_url_parse_data(start_month, start_year, num_of_month, company)
save_to_file(list_data_whole_range,data_file_name)

#######################################

"""


dmc_1997_01_248m_list_data = load_from_file("damacai_1997_01_248_m.txt")
magnum_1996_01_260m_list_data = load_from_file("magnum_1996_01_260_m.txt")
sportstoto_1993_01_296m_list_data = load_from_file("sportstoto_1993_01_296_m.txt")

nr_vs_pcat_1_3drawset_count_p2excel(sportstoto_1993_01_296m_list_data)


"""

dmc_pcat111_nr_3drawset_pattern = nr_3drawSet(dmc_1997_01_248m_list_data, 1,1,1)
dmc_pcat121_nr_3drawset_pattern = nr_3drawSet(dmc_1997_01_248m_list_data, 1,2,1)
dmc_pcat131_nr_3drawset_pattern = nr_3drawSet(dmc_1997_01_248m_list_data, 1,3,1)

dmc_pcat211_nr_3drawset_pattern = nr_3drawSet(dmc_1997_01_248m_list_data, 2,1,1)
dmc_pcat221_nr_3drawset_pattern = nr_3drawSet(dmc_1997_01_248m_list_data, 2,2,1)
dmc_pcat231_nr_3drawset_pattern = nr_3drawSet(dmc_1997_01_248m_list_data, 2,3,1)

dmc_pcat311_nr_3drawset_pattern = nr_3drawSet(dmc_1997_01_248m_list_data, 3,1,1)
dmc_pcat321_nr_3drawset_pattern = nr_3drawSet(dmc_1997_01_248m_list_data, 3,2,1)
dmc_pcat331_nr_3drawset_pattern = nr_3drawSet(dmc_1997_01_248m_list_data, 3,3,1)

dmc_pcat111_nr_counter_list = counter_list_3draw(dmc_pcat111_nr_3drawset_pattern)
dmc_pcat121_nr_counter_list = counter_list_3draw(dmc_pcat121_nr_3drawset_pattern)
dmc_pcat131_nr_counter_list = counter_list_3draw(dmc_pcat131_nr_3drawset_pattern)

dmc_pcat211_nr_counter_list = counter_list_3draw(dmc_pcat211_nr_3drawset_pattern)
dmc_pcat221_nr_counter_list = counter_list_3draw(dmc_pcat221_nr_3drawset_pattern)
dmc_pcat231_nr_counter_list = counter_list_3draw(dmc_pcat231_nr_3drawset_pattern)

dmc_pcat311_nr_counter_list = counter_list_3draw(dmc_pcat311_nr_3drawset_pattern)
dmc_pcat321_nr_counter_list = counter_list_3draw(dmc_pcat321_nr_3drawset_pattern)
dmc_pcat331_nr_counter_list = counter_list_3draw(dmc_pcat331_nr_3drawset_pattern)


port_counter_list_to_xls(dmc_pcat111_nr_counter_list, "111")
port_counter_list_to_xls(dmc_pcat121_nr_counter_list, "121")
port_counter_list_to_xls(dmc_pcat131_nr_counter_list, "131")

port_counter_list_to_xls(dmc_pcat211_nr_counter_list, "211")
port_counter_list_to_xls(dmc_pcat221_nr_counter_list, "221")
port_counter_list_to_xls(dmc_pcat231_nr_counter_list, "231")

port_counter_list_to_xls(dmc_pcat311_nr_counter_list, "311")
port_counter_list_to_xls(dmc_pcat321_nr_counter_list, "321")
port_counter_list_to_xls(dmc_pcat331_nr_counter_list, "331")



#save_to_file(built_url_parse_data()) ## step 1, 2, 3
#data_3draw_set = p1_nr_3drawSet()
#count_3draw_set = counter_list_3draw(data_3draw_set)

"""

print
print
print("#####################################")
print ("#####    END OF EXECUTION       #####")
print("#####################################")