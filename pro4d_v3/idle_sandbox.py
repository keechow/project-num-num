from  get_result_frm_webpage import build_url_search_term
from  get_result_frm_webpage import parse_data_from_url
from openpyxl import load_workbook

import num_analysis

import json

import collections

## save list_data_whole_range into  file for future use
def save_to_file(data):
	with open("data_file.txt","w") as outputfile:
		json.dump(data, outputfile)

def load_from_file():
	return_list_data = []
	with open("data_file.txt") as inputfile:
		json_load_from_file = json.load(inputfile)
		for each_draw in json_load_from_file:
			str_data = [s.encode('utf-8') for s in each_draw]
			return_list_data.append(str_data)
	return return_list_data

def built_url_parse_data():
	str_data_collection_start_month = "01"
	str_data_collection_start_year = "1997"
	str_data_collection_num_of_month = "240"
	str_data_collection_company = "damacai"

	list_search_url = build_url_search_term(str_data_collection_start_month, str_data_collection_start_year, str_data_collection_num_of_month, str_data_collection_company)
		## containing a list of URL to be used to request data from Internet

	list_data_whole_range = []
		## list for containing all parsed data

	for each_month_data_url in list_search_url:
		monthly_data = parse_data_from_url(each_month_data_url)
		for each_draw_data in monthly_data:
			list_data_whole_range.append(each_draw_data)

	return list_data_whole_range

def p1_nr_3drawSet():
	data_from_file = load_from_file()
	max_loop_count = len(data_from_file) - 3
	counter = 0

	list_p1_nr_3drawSet = []

	while counter <= max_loop_count:
		num_range1 = data_from_file[counter][1][0]
		num_range2 = data_from_file[counter + 1][1][0]
		num_range3 = data_from_file[counter + 2][1][0]

		list_p1_nr_3drawSet.append(num_range1+num_range2+num_range3)
		counter += 1

	return list_p1_nr_3drawSet

def p1_nr_4drawSet():
	data_from_file = load_from_file()
	max_loop_count = len(data_from_file) - 4
	counter = 0

	list_p1_nr_4drawSet = []

	while counter <= max_loop_count:
		num_range1 = data_from_file[counter][1][0]
		num_range2 = data_from_file[counter + 1][1][0]
		num_range3 = data_from_file[counter + 2][1][0]
		num_range4 = data_from_file[counter + 3][1][0]
		list_p1_nr_4drawSet.append(num_range1+num_range2+num_range3+num_range4)
		counter += 1

	return list_p1_nr_4drawSet

	## param: a list containing num range extracted from 3 consecutive draws
	## count each occurrence of 3draw set pattern - 000 ~ 999
	## 1) Create a counter list containing 100 integer 0
	## 2) Iterate through the input list and add count to respective element in counter_list



list_nr_3drawSet = p1_nr_3drawSet()


def counter_list(drawSet_data):
	return_counter_list = []
	for each in range(1000):
		return_counter_list.append(0)

	for each in drawSet_data:
		int_each = int(each)
		count =return_counter_list[int_each] + 1
		return_counter_list[int_each] = count

	return return_counter_list


def counter_list_processor(list_counter_data):
	# total of all count in counter list
	total_sum = 0
	for each in counter_list:
		total_sum += each

	# average of all count
	avg_count = int(total_sum / (len(counter_list)))

	#calculate median
	sorted_counter_list = sorted(counter_list)
	median_pos = int(len(counter_list)/2)
	median_count = sorted_counter_list[median_pos]

	#identify numrange 3Draw set pattern that gives avg and median count
	list_avg_pattern_pos = []
	list_median_pattern_pos = []

	for each_count in counter_list:
		if each_count == avg_count:
			list_avg_pattern_pos.append(counter_list.index(each_count))
		if each_count == median_count:
			list_median_pattern_pos.append((counter_list.index(each_count)))

	#print(counter_list)

	# time to process the count in counter_list
	# 1) Highest count value
	# 2) Lowest count value

	count_max_value = max(counter_list)
	count_min_value = min(counter_list)

def port_counter_list_to_xls(list_counter):
	# params: counter list
	# objective: port the counter list into xls spreadsheet for easy visualization and further studies

	#1. load xls workbook
	my_wb = load_workbook("list_3drawset_pattern.xlsx")
	#2. load worksheet in workbook
	my_ws = my_wb.get_sheet_by_name("01")

	drawset_counter = 0
	while drawset_counter < len(list_counter):
	#3. set initial cell address to port in data
		current_row = drawset_counter + 1
		drawset__column = 1
		count_column = 2
		str_drawset_pattern = str(drawset_counter)
		while (len(str_drawset_pattern) < 3):
			str_drawset_pattern = "0" + str_drawset_pattern
		str_drawset_count = list_counter[drawset_counter]

		my_ws.cell(row = current_row, column = drawset__column).value = str_drawset_pattern
		my_ws.cell(row = current_row, column = count_column).value = str_drawset_count
		drawset_counter += 1
	my_wb.save("list_3drawset_pattern.xlsx")

port_counter_list_to_xls(counter_list(list_nr_3drawSet))
