__author__ = 'keechow'

from  get_result_frm_webpage import build_url_search_term
from  get_result_frm_webpage import parse_data_from_url
from openpyxl import load_workbook
from calendar import weekday

import num_analysis

import json

## save list_data_whole_range into  file for future use
def save_to_file(data, file_name):
	with open(file_name,"w") as outputfile:
		json.dump(data, outputfile)

def load_from_file(file_name):
	return_list_data = []
	with open(file_name) as inputfile:
		json_load_from_file = json.load(inputfile)
		for each_draw in json_load_from_file:
			str_data = [s.encode('utf-8') for s in each_draw]
			return_list_data.append(str_data)
	return return_list_data

def built_url_parse_data(s_month, s_year, num_month, company):
	str_data_collection_start_month = s_month
	str_data_collection_start_year = start_year
	str_data_collection_num_of_month = num_month
	str_data_collection_company = company

	list_search_url = build_url_search_term(str_data_collection_start_month, str_data_collection_start_year, str_data_collection_num_of_month, str_data_collection_company)
		## containing a list of URL to be used to request data from Internet

	list_data_whole_range = []
		## list for containing all parsed data

	for each_month_data_url in list_search_url:
		print each_month_data_url
		monthly_data = parse_data_from_url(each_month_data_url)
		for each_draw_data in monthly_data:
			list_data_whole_range.append(each_draw_data)

	return list_data_whole_range

def remove_duplicate (list_data):
	# removes duplicates in list_data
	clean_list = []
	for each in list_data:
		if not each in clean_list:
			clean_list.append(each)
	return(sorted(clean_list))

def counter_list_3draw(drawSet_data):
	# params: list of data containing draw set pattern
	# count each draw set pattern
	# return a list containing count for each draw set pattern in respective position
	# e.g. element [203] correspond to draw set pattern 203
	return_counter_list = []
	for each in range(1000):
		return_counter_list.append(0)

	for each in drawSet_data:
		int_each = int(each)
		count = return_counter_list[int_each] + 1
		return_counter_list[int_each] = count

	return return_counter_list

def port_counter_list_to_xls(list_counter, s_name):
	# params: counter list
	# objective: port the counter list into xls spreadsheet for easy visualization and further studies

	#1. load xls workbook
	my_wb =  load_workbook("list_3drawset_pattern.xlsx")

	#2. load worksheet in workbook
	sheet_names = my_wb.sheetnames

	if s_name in sheet_names:
		my_ws = my_wb.get_sheet_by_name(s_name)
	else:
		my_ws = my_wb.create_sheet(title=s_name)

	drawset_counter = 0
	while drawset_counter < len(list_counter):
	#3. set initial cell address to port in data
		current_row = drawset_counter + 1
		drawset__column = 1
		count_column = 2
		str_drawset_pattern = str(drawset_counter)
		while (len(str_drawset_pattern) < 3):
			str_drawset_pattern = "0" + str_drawset_pattern
		str_drawset_count = list_counter[drawset_counter]

		my_ws.cell(row = current_row, column = drawset__column).value = str_drawset_pattern
		my_ws.cell(row = current_row, column = count_column).value = str_drawset_count
		drawset_counter += 1
	my_wb.save("list_3drawset_pattern.xlsx")

def identify_zero_count_pattern(list_data):
	# return 3 draw set pattern that has zero occurrence
	list_1K_pattern = []
	list_zero_count_pattern = []
	for each_num in range(1000):
		str_num = str(each_num)
		while (len(str_num) < 3):
			str_num = "0" + str_num
		list_1K_pattern.append(str_num)

	for each in list_1K_pattern:
		if not each in list_data:
			list_zero_count_pattern.append(each)

	return list_zero_count_pattern

def gen_data_list_by_day_of_week(list_data):
	# params: data list from txt file
	# obj: separate draw data into Wed, Sat, Sun, Special draw day
	# return: a list containing 4 list element, [wed, sat, sun, other]
	list_wed_draw = []
	list_sat_draw = []
	list_sun_draw = []
	list_other_draw = []

	#list to store draw result for Wed, Sat, Sun and other day

	dict_draw_day_list = {2:list_wed_draw, 5:list_sat_draw, 6:list_sun_draw}

	for each_data_set in list_data:
		# [["19970101", "4102", "5257", "0061"],....
		# extract year, month, day from data and use calendar.weekday() to determine day of the week

		year = int(each_data_set[0][:4])
		month =  int(each_data_set[0][4:6])
		day =  int(each_data_set[0][6:])

		day_of_week = weekday(year,month,day)

		if day_of_week in dict_draw_day_list:
			dict_draw_day_list[day_of_week].append(each_data_set)
		else:
			list_other_draw.append(each_data_set)

	return [list_wed_draw, list_sat_draw, list_sun_draw, list_other_draw]

def nr_drawSet(list_data, list_pcat):
	# params: list_data
	# params: list_pcat - type list, containing pcat for num range

	num_of_draw = len(list_pcat)
	max_loop_count = len(list_data) - num_of_draw
	loop_count = 0
	return_list_nr_drawset = []

	while loop_count <= max_loop_count:
		drawset_pattern = ""
		pcat_count = 0
		while pcat_count < num_of_draw:
			pcat = list_pcat[pcat_count]
			pattern = list_data[loop_count + pcat_count][pcat][0]
			drawset_pattern += pattern
			pcat_count += 1
		return_list_nr_drawset.append(drawset_pattern)
		loop_count +=  1

	return return_list_nr_drawset

def nr_3drawSet(list_data, p_cat0, p_cat1, p_cat3):
	## 1. get data from input
	## 2. Extract num range for 3 consecutive draws and save into a list
	## 3. Return the list
	max_loop_count = len(list_data) - 3
	counter = 0

	list_p1_nr_3drawSet = []

	while counter <= max_loop_count:
		num_range1 = list_data[counter][p_cat0][0]
		num_range2 = list_data[counter + 1][p_cat1][0]
		num_range3 = list_data[counter + 2][p_cat3][0]

		list_p1_nr_3drawSet.append(num_range1+num_range2+num_range3)
		counter += 1

	return list_p1_nr_3drawSet

def nr_4drawSet(list_data, p_cat0, p_cat1, p_cat2, p_cat3):
	## 1. Get data from input
	## 2. Extract num range for 4 consecutive draws and save into a list
	## 3. Return the list

	max_loop_count = len(list_data) - 4
	counter = 0

	list_p1_nr_4drawSet = []

	while counter <= max_loop_count:
		num_range1 = list_data[counter][p_cat0][0]
		num_range2 = list_data[counter + 1][p_cat1][0]
		num_range3 = list_data[counter + 2][p_cat2][0]
		num_range4 = list_data[counter + 3][p_cat3][0]
		list_p1_nr_4drawSet.append(num_range1+num_range2+num_range3+num_range4)
		counter += 1

	return list_p1_nr_4drawSet

def nr_3drawSet_dmc3d(list_data, p_cat0, p_cat1, p_cat3):
	## 1. get data from input
	## 2. Extract num range for 3 consecutive draws and save into a list
	## 3. Return the list
	max_loop_count = len(list_data) - 3
	counter = 0

	list_p1_nr_3drawSet = []

	while counter <= max_loop_count:
		num_range1 = list_data[counter][p_cat0][1]
		num_range2 = list_data[counter + 1][p_cat1][1]
		num_range3 = list_data[counter + 2][p_cat3][1]

		list_p1_nr_3drawSet.append(num_range1+num_range2+num_range3)
		counter += 1

	return list_p1_nr_3drawSet

def nr_vs_pcat_1_3drawset_count_p2excel(list_data):
	# params: data list from txt file
	# 1) extract 3 draw set pattern (nr vs pcat)
	# 2) count occurrence for each 3draw set pattern
	# 3) port the counter list to excel for future analysis
	# 4) pcat pattern = 111,121,131,  211,221,231, 311,321,331

	for pcat0 in range(1,4):
		for pcat1 in range(1,4):
			nr_3drawset_pattern = nr_3drawSet(list_data, pcat0, pcat1, 1)
			nr_counter_list = counter_list_3draw(nr_3drawset_pattern)
			sheet_name = str(pcat0) + str(pcat1) + "1"
			port_counter_list_to_xls(nr_counter_list, sheet_name)

def nr_vs_pcat_1_3drawset_count_p2excel_dmc3d(list_data):
	# params: data list from txt file
	# 1) extract 3 draw set pattern (nr vs pcat)
	# 2) count occurrence for each 3draw set pattern
	# 3) port the counter list to excel for future analysis
	# 4) pcat pattern = 111,121,131,  211,221,231, 311,321,331

	for pcat0 in range(1,4):
		for pcat1 in range(1,4):
			nr_3drawset_pattern = nr_3drawSet_dmc3d(list_data, pcat0, pcat1, 2)
			nr_counter_list = counter_list_3draw(nr_3drawset_pattern)
			sheet_name = str(pcat0) + str(pcat1) + "2"
			port_counter_list_to_xls(nr_counter_list, sheet_name)


##########################################
##                                  RUN  CODE                                 ##
##########################################

"""
### Run once to get data from Internet and save into file ###

start_month = "01"
start_year = "2011"
num_of_month = "81"
company = "magnum"

data_file_name = company + "_" + start_year + "_" + start_month + "_" + num_of_month + "_m.txt"

list_data_whole_range = built_url_parse_data(start_month, start_year, num_of_month, company)
save_to_file(list_data_whole_range,data_file_name)

#######################################

"""

"""
## load data from txt file
## separate draw data in accordance to day of week
## save DOW data into txt file

dmc_1997_01_248m_list_data = load_from_file("damacai_1997_01_248_m.txt")
magnum_1996_01_260m_list_data = load_from_file("magnum_1996_01_260_m.txt")
sportstoto_1993_01_296m_list_data = load_from_file("sportstoto_1993_01_296_m.txt")

dmc_1997_01_248m_list_data_DOW = gen_data_list_by_day_of_week(dmc_1997_01_248m_list_data)
magnum_1996_01_260m_list_data_DOW =gen_data_list_by_day_of_week(magnum_1996_01_260m_list_data)
sportstoto_1993_01_296m_list_data_DOW =gen_data_list_by_day_of_week(sportstoto_1993_01_296m_list_data)


save_to_file(magnum_1996_01_260m_list_data,"magnum_1996_01_260m_list_data.txt")
save_to_file(sportstoto_1993_01_296m_list_data,"sportstoto_1993_01_296m_list_data.txt")
save_to_file(dmc_1997_01_248m_list_data_DOW, "dmc_1997_01_248m_list_data_DOW.txt")
save_to_file(magnum_1996_01_260m_list_data_DOW, "magnum_1996_01_260m_list_data_DOW.txt")
save_to_file(sportstoto_1993_01_296m_list_data_DOW, "sportstoto_1993_01_296m_list_data_DOW.txt")
"""

#"""
# This is for DOW data analysis
# Run 3 drawset vs nr analysis and port to excel for charting
# data_list + pcat0,pcat1,pcat2 --> nr3setdraw --> ctr_list_3draw --> port_ctr_list


dmc_1997_01_248m_list_data = load_from_file("damacai_1997_01_248_m.txt")

dmc_1997_01_248m_list_data_DOW = gen_data_list_by_day_of_week(dmc_1997_01_248m_list_data)

nr_vs_pcat_1_3drawset_count_p2excel_dmc3d(dmc_1997_01_248m_list_data)



"""
magnum_1996_01_260m_list_data_DOW =gen_data_list_by_day_of_week(magnum_1996_01_260m_list_data)
sportstoto_1993_01_296m_list_data_DOW =gen_data_list_by_day_of_week(sportstoto_1993_01_296m_list_data)

dmc_1997_01_248m_list_data_DOW_wed = dmc_1997_01_248m_list_data_DOW[0]
dmc_1997_01_248m_list_data_DOW_sat = dmc_1997_01_248m_list_data_DOW[1]
dmc_1997_01_248m_list_data_DOW_sun = dmc_1997_01_248m_list_data_DOW[2]
dmc_1997_01_248m_list_data_DOW_oth = dmc_1997_01_248m_list_data_DOW[3]

sportstoto_1993_01_296m_list_data_DOW_wed = sportstoto_1993_01_296m_list_data_DOW[0]
sportstoto_1993_01_296m_list_data_DOW_sat = sportstoto_1993_01_296m_list_data_DOW[1]
sportstoto_1993_01_296m_list_data_DOW_sun = sportstoto_1993_01_296m_list_data_DOW[2]
sportstoto_1993_01_296m_list_data_DOW_oth = sportstoto_1993_01_296m_list_data_DOW[3]

magnum_1996_01_260m_list_data_DOW_wed = magnum_1996_01_260m_list_data_DOW[0]
magnum_1996_01_260m_list_data_DOW_sat = magnum_1996_01_260m_list_data_DOW[1]
magnum_1996_01_260m_list_data_DOW_sun = magnum_1996_01_260m_list_data_DOW[2]
magnum_1996_01_260m_list_data_DOW_oth = magnum_1996_01_260m_list_data_DOW[3]


nr_vs_pcat_1_3drawset_count_p2excel(magnum_1996_01_260m_list_data_DOW_wed)
"""
#"""
print
print
print("#####################################")
print ("#####    END OF EXECUTION       #####")
print("#####################################")